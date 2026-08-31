# -*- coding: utf-8 -*-
import openpyxl

path = r"C:\Users\Administrator\Desktop\成品定点定位看板.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb['三栋五楼成品']
print("三栋五楼成品 max_row=", ws.max_row, "max_column=", ws.max_column)
for r in range(1, ws.max_row + 1):
    rowvals = []
    for c in range(1, min(ws.max_column, 8) + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            rowvals.append(f"C{c}={str(v).strip()!r}")
    if rowvals:
        print(f"row{r}: " + " | ".join(rowvals))
wb.close()
