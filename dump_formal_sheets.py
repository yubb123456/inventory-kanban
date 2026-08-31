# -*- coding: utf-8 -*-
import openpyxl

path = r"C:\Users\Administrator\Desktop\成品定点定位看板.xlsx"
wb = openpyxl.load_workbook(path, read_only=True)
print("Sheet 数量:", len(wb.sheetnames))
print("Sheet 列表:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    cells = []
    for r in range(1, min(ws.max_row, 15) + 1):
        rowvals = []
        for c in range(1, min(ws.max_column, 8) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                rowvals.append(f"C{c}={str(v).strip()!r}")
        if rowvals:
            cells.append(f"  row{r}: " + " | ".join(rowvals))
    print(f"\n=== {name} (max_row={ws.max_row}) ===")
    for line in cells:
        print(line)
wb.close()
