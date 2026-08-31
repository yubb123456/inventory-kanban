#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
成品定点定位看板 数据解析脚本
==============================
将《成品定点定位看板.xlsx》解析为看板数据 src/data/kanban.json

用法:
    python scripts/parse_excel.py <Excel路径> [输出JSON路径]

示例:
    python scripts/parse_excel.py "D:/库存/成品定点定位看板.xlsx"
    python scripts/parse_excel.py "D:/库存/成品定点定位看板.xlsx" src/data/kanban.json

说明:
    - 每个 Sheet 视为一个仓库区域
    - 每 2 列为一组储位（编码 + 规格型号），最多 4 组（列 1-8）
    - 货架标签行（如 "成品一区1-1"、"阁楼A货架"）定义储位归属
    - 子标签行（左/右/前/后、A1层）进一步细分储位
    - 表头行（编码/规格型号）与区域标题行会被自动识别跳过
"""
import json
import re
import sys
from collections import OrderedDict

# 兼容 openpyxl 未安装的情况
try:
    import openpyxl
except ImportError:
    print("错误：未安装 openpyxl，请先执行 pip install openpyxl")
    sys.exit(1)

CODE_RE = re.compile(r'^\d{2}\.\d{2}\.\d{2}\.\d{4}')


def get_cell(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_sheet(ws):
    name = ws.title
    rows_data = []
    for r in range(1, ws.max_row + 1):
        rows_data.append([ws.cell(r, c).value for c in range(1, 9)])

    zone_title = name
    racks = []                      # [{name, slot_map:{sub: slot}}]
    rack_by_name = {}
    cur_rack_ctx = {}               # col_idx -> {"rack": rack, "sub": str}

    for vals in rows_data:
        cells = [get_cell(v) for v in vals]
        nonempty = [(i, c) for i, c in enumerate(cells) if c is not None]
        if not nonempty:
            continue
        texts = [c for _, c in nonempty]

        # 总看板标题
        if any('定点定位看板' in c for c in texts):
            continue

        # 表头行
        if any('编码' in c for c in texts):
            continue

        # 区域标题行：单标签，等于 sheet 名或以 区/楼/仓 结尾且不含货架/矩阵特征词
        is_zone_title = (len(nonempty) == 1 and (
            texts[0] == name or
            (texts[0].endswith(('区', '楼', '仓')) and
             not any(k in texts[0] for k in ('货架', '矩阵', '编号')))
        ))
        if is_zone_title:
            zone_title = texts[0]
            continue

        # 数据行：含编码
        has_code = any(CODE_RE.match(c) for c in texts)
        if has_code:
            for i, c in enumerate(cells):
                if i in (0, 2, 4, 6) and c and CODE_RE.match(c):
                    ci = i // 2
                    spec = cells[i + 1] if i + 1 < len(cells) else None
                    ctx = cur_rack_ctx.get(ci)
                    if ctx is None:
                        rack = rack_by_name.get(zone_title)
                        if rack is None:
                            rack = {"name": zone_title, "slot_map": {}}
                            rack_by_name[zone_title] = rack
                            racks.append(rack)
                        ctx = {"rack": rack, "sub": ""}
                        cur_rack_ctx[ci] = ctx
                    rack = ctx["rack"]
                    sub = ctx.get("sub") or ""
                    slot = rack["slot_map"].get(sub)
                    if slot is None:
                        slot = {"sub": sub, "items": []}
                        rack["slot_map"][sub] = slot
                    slot["items"].append({"code": c, "spec": spec})
            continue

        # 子标签行：所有标签都是 左/右/前/后 或以"层"结尾
        is_sublabel = texts and all(
            (t in ('左', '右', '前', '后')) or t.endswith('层') for t in texts
        )
        if is_sublabel:
            label_cols = {i: c for i, c in enumerate(cells) if i in (0, 2, 4, 6) and c}
            sorted_idx = sorted(label_cols.keys())
            for gi in range(4):
                owner = None
                for si in sorted_idx:
                    if si // 2 <= gi:
                        owner = si
                if owner is not None and gi in cur_rack_ctx:
                    cur_rack_ctx[gi]["sub"] = label_cols[owner]
            continue

        # 货架标签行
        label_cols = {i: c for i, c in enumerate(cells) if i in (0, 2, 4, 6) and c}
        if not label_cols:
            continue
        sorted_idx = sorted(label_cols.keys())
        new_rack_ctx = {}
        for gi in range(4):
            owner = None
            for si in sorted_idx:
                if si // 2 <= gi:
                    owner = si
            rname = label_cols.get(owner) if owner is not None else None
            if rname:
                rack = rack_by_name.get(rname)
                if rack is None:
                    rack = {"name": rname, "slot_map": {}}
                    rack_by_name[rname] = rack
                    racks.append(rack)
                new_rack_ctx[gi] = {"rack": rack, "sub": ""}
        cur_rack_ctx = new_rack_ctx

    zone_racks = []
    for rack in racks:
        slots = [{"sub": s, "items": slot["items"]} for s, slot in rack["slot_map"].items()]
        slots.sort(key=lambda x: (0 if x["sub"] == "" else 1, x["sub"]))
        zone_racks.append({"name": rack["name"], "slots": slots})
    return {"name": name, "title": zone_title, "racks": zone_racks}


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    excel_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "src/data/kanban.json"

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    zones = []
    for sn in wb.sheetnames:
        if wb[sn].max_row <= 1:
            continue
        z = parse_sheet(wb[sn])
        if z["racks"]:
            zones.append(z)

    total_items = sum(
        len(s["items"]) for z in zones for rk in z["racks"] for s in rk["slots"]
    )

    payload = {"zones": zones, "totalItems": total_items}
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    print(f"解析完成：{len(zones)} 个区域，{total_items} 条在库 SKU")
    print(f"输出文件：{out_path}")


if __name__ == "__main__":
    main()
